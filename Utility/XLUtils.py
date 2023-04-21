import openpyxl
from openpyxl import load_workbook


class XLUtils:

    def getRowCount(self, file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheetNames = workbook.sheetnames
        for name in sheetNames:
            print(name)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(self, sssssssssssssssssaaaaAAAAAAAAAAAAAAAAAAAAAAQwq 5C4EFRVD Dfile, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)

    def readData(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)






